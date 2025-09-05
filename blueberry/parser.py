from typing import Any, TypeVar, Union
from datetime import time as datetime_time
import warnings

from pydantic import BaseModel
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from .models import (
    AppendOnlyModel,
    DeleteModel,
    LongTaskModel,
    ProgressModel,
    ShortTaskModel,
    WorktimeModel,
    PickerModel,
)

# openpyxl
warnings.simplefilter(action="ignore", category=UserWarning)

T = TypeVar("T")


class Data(BaseModel):
    长期任务: list[Union[LongTaskModel, DeleteModel]]
    长期进度: list[ProgressModel]
    短期任务: list[Union[ShortTaskModel, DeleteModel]]
    工作时段: list[WorktimeModel] = [
        WorktimeModel(开始=datetime_time(hour=0), 结束=datetime_time(hour=0))
    ]
    选择排序偏好: list[PickerModel] = []


def parse_table(table: Worksheet) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    keys: list[str] = []
    for row in table.iter_rows(min_row=1):
        for cell in row:
            keys.append(str(cell.value))
    for row in table.iter_rows(min_row=2):
        row_: dict[str, Any] = {}
        for i, cell in enumerate(row):
            if cell.value is None:
                continue
            row_[keys[i]] = cell.value
        if row_:
            rows.append(row_)
    return rows


def parse_model_table(table: Worksheet, datatype: type[T]) -> list[T]:
    parsed_table = parse_table(table)
    ret: list[T] = []
    for x in parsed_table:
        ret.append(datatype(**x))
    return ret


def parse_append_only_table(
    table: Worksheet, datatype: type[AppendOnlyModel]
) -> list[Union[AppendOnlyModel, DeleteModel]]:
    parsed_table = parse_table(table)
    ret: list[Union[AppendOnlyModel, DeleteModel]] = []
    for x in parsed_table:
        if set(x.keys()) == {"名称", "时间"}:
            ret.append(DeleteModel(**x))
        else:
            ret.append(datatype(**x))
    return ret


def load_data(workbook: str) -> Data:
    wb = load_workbook(workbook)
    长期任务 = parse_append_only_table(wb["长期任务"], LongTaskModel)
    长期进度 = parse_model_table(wb["长期进度"], ProgressModel)
    短期任务 = parse_append_only_table(wb["短期任务"], ShortTaskModel)
    工作时段 = [WorktimeModel(开始=datetime_time(hour=0), 结束=datetime_time(hour=0))]
    选择排序偏好: list[PickerModel] = []
    if "工作时段" in wb.sheetnames:
        工作时段 = parse_model_table(wb["工作时段"], WorktimeModel)
    if "选择排序偏好" in wb.sheetnames:
        选择排序偏好 = parse_model_table(wb["选择排序偏好"], PickerModel)
    return Data(
        长期任务=长期任务,
        长期进度=长期进度,
        短期任务=短期任务,
        工作时段=工作时段,
        选择排序偏好=选择排序偏好,
    )
