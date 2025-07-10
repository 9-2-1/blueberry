from typing import Literal, Any, TypeVar, Generic, Callable, overload
from dataclasses import dataclass
from datetime import datetime, timedelta, time as datetime_time
from textwrap import indent, wrap
import math
import sys
import os
import warnings
import aiohttp.web
import argparse

from pydantic import BaseModel, ConfigDict
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
import wcwidth  # type: ignore

# 用于tabulate的宽度计算
from tabulate import tabulate

warnings.simplefilter(action="ignore", category=UserWarning)

# 8小时(时长) * 80%(工作-休息比) ≈ 6.5小时
推荐用时 = timedelta(hours=6.5)


class AppendOnly(BaseModel):
    """
    记录表中共有的两列，一列代表记录的名称，一列代表记录添加、更新或删除的时间。
    """

    名称: str
    时间: datetime


class DeleteModel(AppendOnly):
    """
    代表“删除”的记录行。
    在“时间”删除名为“名称”的行。
    """

    名称: str
    时间: datetime


class TaskModel(AppendOnly):
    model_config = ConfigDict(extra="forbid")
    名称: str
    时间: datetime
    # ---
    标题: str
    描述: str | None = None
    开始: datetime
    结束: datetime
    总数: float | None = None


class ProgressModel(BaseModel):
    model_config = ConfigDict(extra="forbid")
    时间: datetime
    名称: str
    # ---
    进度: float = 0.0
    用时: timedelta = timedelta(0)
    描述: str | None = None


class StatusModel(AppendOnly):
    model_config = ConfigDict(extra="forbid")
    名称: str
    时间: datetime
    # ---
    标题: str
    描述: str | None = None
    点数: int | None = None  # 点数为None意味着这是一个“取消之前状态”的记录。
    开始: datetime | None = None
    结束: datetime | None = None


class TodoModel(AppendOnly):
    model_config = ConfigDict(extra="forbid")
    名称: str
    时间: datetime
    # ---
    标记: Literal["+", "*", "-"] | None = None  # + 未开始，* 继续，- 等待
    标题: str
    描述: str | None = None
    点数: int = 0
    开始: datetime | None = None
    结束: datetime | None = None
    完成: datetime | None = None


class HintModel(BaseModel):
    model_config = ConfigDict(extra="forbid")
    时间: datetime
    # ---
    标题: str
    描述: str | None = None


class WorktimeModel(BaseModel):
    开始: datetime_time
    结束: datetime_time


AppendOnlyModel = TypeVar("AppendOnlyModel", bound=AppendOnly)
T = TypeVar("T")


def parse_table(table: Worksheet) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    keys: list[str] = []
    for row in table.iter_rows(min_row=1):
        for cell in row:
            keys.append(str(cell.value))
    for row in table.iter_rows(min_row=2):
        row_: dict[str, Any] = {}
        for i, cell in enumerate(row):
            if cell.value is None:
                continue
            row_[keys[i]] = cell.value
        if row_:
            rows.append(row_)
    return rows


def parse_model_table(table: Worksheet, datatype: type[T]) -> list[T]:
    parsed_table = parse_table(table)
    ret: list[T] = []
    for x in parsed_table:
        ret.append(datatype(**x))
    return ret


def parse_append_only_table(
    table: Worksheet, datatype: type[AppendOnlyModel]
) -> list[AppendOnlyModel | DeleteModel]:
    parsed_table = parse_table(table)
    ret: list[AppendOnlyModel | DeleteModel] = []
    for x in parsed_table:
        if set(x.keys()) == {"名称", "时间"}:
            ret.append(DeleteModel(**x))
        else:
            ret.append(datatype(**x))
    return ret


class Data(BaseModel):
    任务: list[TaskModel | DeleteModel]
    进度: list[ProgressModel]
    状态: list[StatusModel | DeleteModel]
    待办事项: list[TodoModel | DeleteModel]
    提示: list[HintModel]
    工作时段: list[WorktimeModel] = [
        WorktimeModel(开始=datetime_time(hour=0), 结束=datetime_time(hour=0))
    ]


class State(BaseModel):
    任务: dict[str, TaskModel]
    进度: dict[str, list[ProgressModel]]
    状态: dict[str, StatusModel]
    待办事项: dict[str, TodoModel]
    提示: list[HintModel]
    工作时段: list[WorktimeModel]


def load_data(workbook: str) -> Data:
    wb = load_workbook(workbook)
    任务 = parse_append_only_table(wb["任务"], TaskModel)
    进度 = parse_model_table(wb["进度"], ProgressModel)
    状态 = parse_append_only_table(wb["状态"], StatusModel)
    待办事项 = parse_append_only_table(wb["待办事项"], TodoModel)
    提示 = parse_model_table(wb["提示"], HintModel)
    工作时段 = [WorktimeModel(开始=datetime_time(hour=0), 结束=datetime_time(hour=0))]
    if "工作时段" in wb.sheetnames:
        工作时段 = parse_model_table(wb["工作时段"], WorktimeModel)
    return Data(
        状态=状态, 任务=任务, 进度=进度, 待办事项=待办事项, 提示=提示, 工作时段=工作时段
    )


def collect_lines(
    lines: list[AppendOnlyModel | DeleteModel], now_time: datetime
) -> dict[str, AppendOnlyModel]:
    lines = sorted(lines, key=lambda x: x.时间)
    state: dict[str, AppendOnlyModel] = {}
    for line in lines:
        if line.时间 >= now_time:
            continue
        if isinstance(line, DeleteModel):
            del state[line.名称]
        else:
            state[line.名称] = line
    return state


def collect_progress(
    progress: list[ProgressModel], now_time: datetime
) -> dict[str, list[ProgressModel]]:
    progress = sorted(progress, key=lambda x: x.时间)
    state: dict[str, list[ProgressModel]] = {}  # list of progress nodes
    for line in progress:
        if line.时间 >= now_time:
            continue
        if line.名称 in state:
            state[line.名称].append(line)
        else:
            state[line.名称] = [line]
    return state


def collect_hints(lines: list[HintModel], now_time: datetime) -> list[HintModel]:
    lines = sorted(lines, key=lambda x: x.时间)
    return [x for x in lines if x.时间 <= now_time]


def collect_state(data: Data, now_time: datetime) -> State:
    return State(
        任务=collect_lines(data.任务, now_time),
        进度=collect_progress(data.进度, now_time),
        状态=collect_lines(data.状态, now_time),
        待办事项=collect_lines(data.待办事项, now_time),
        提示=collect_hints(data.提示, now_time),
        工作时段=data.工作时段,
    )


@dataclass
class TaskSpeed:
    速度: float  # per hour
    每日用时: timedelta


@dataclass
class TaskEstimate:
    预计完成时间: timedelta
    预计可用时间: timedelta
    差距: timedelta


@dataclass
class TaskStats:
    进度: float
    用时: timedelta
    标记: Literal["*", "-", "=", "!"]
    进度描述: str | None = None
    # * 进行中 - 等待开始 = 已完成 ! 已超时
    速度: TaskSpeed | None = None  # 开始的任务才能计算速度
    预计: TaskEstimate | None = None  # 开始但未完成的任务才能估计完成时间
    点数: int | None = (
        None  # 点数。到时未开始: -100 * 延后天数。已开始：100 * (时间差距 / 每日用时)。已结束：100 * 剩余天数
    )


@dataclass
class StateStats:
    Goldie点数: int
    任务点数: int
    任务统计: dict[str, TaskStats]
    状态点数: int
    状态生效: list[str]
    其他任务点数: int
    其他任务生效: list[str]
    总每日用时: timedelta


def workday_time(time: datetime, worktime: list[WorktimeModel]) -> timedelta:
    # 计算某个时间点当天已经过了多少工作时间。
    delta = timedelta(0)
    for wt in worktime:
        wt_begin = datetime.combine(time, wt.开始)
        wt_end = datetime.combine(time, wt.结束)
        # 跨天的情况
        if wt_end <= wt_begin:
            wt_end += timedelta(days=1)
        if time < wt_begin:
            pass
        elif time < wt_end:
            delta += time - wt_begin
        else:
            delta += wt_end - wt_begin
    return delta


def workdays(begin: datetime, end: datetime, worktime: list[WorktimeModel]) -> float:
    # 计算两个时间之间的工作时长
    # 注意，这里将一天工作时长视为“一天”。
    # 如果实际的工作时长是 3 小时，而一天的工作时间和是 6 小时，那么就会得到 0.5 天

    # 为了避免过多的分类讨论，使用定数法
    begin_date = begin.date()
    end_date = end.date()

    begin_time = begin.time()
    end_time = end.time()

    datediff = (end_date - begin_date) / timedelta(days=1)

    begin_time_day = workday_time(begin, worktime)
    end_time_day = workday_time(end, worktime)
    total_time_day = timedelta(0)

    for wt in worktime:
        wt_begin = datetime.combine(end, wt.开始)
        wt_end = datetime.combine(end, wt.结束)
        if wt_end <= wt_begin:
            wt_end += timedelta(days=1)
        total_time_day += wt_end - wt_begin

    timediff = (end_time_day - begin_time_day) / total_time_day
    return datediff + timediff


def calculate_speed(
    progress: list[ProgressModel],
    begin_time: datetime,
    now_time: datetime,
    worktime: list[WorktimeModel],
) -> TaskSpeed | None:
    # 近期记录
    MIN_TIME = timedelta(hours=2)
    MIN_TIMESPAN = 3
    new_time = now_time
    tot_time = timedelta(0)
    new_progress = progress[-1].进度
    found = False
    count = 0
    # 这里需要注意时间的计算方式：
    # 进度记录描述的是 “花费‘用时’时间后，在‘时间’让进度达到了‘进度’”。
    # 选择“开始的记录”后“开始的记录”本身的时间不包含在总时间内（那是起点）
    for p in reversed(progress):
        old_time = p.时间
        if (
            tot_time > MIN_TIME
            and workdays(old_time, new_time, worktime) > MIN_TIMESPAN
        ):
            found = True
            old_progress = p.进度
            break
        tot_time += p.用时
        if p.用时 != timedelta(0):
            count += 1
    if not found:
        if begin_time < old_time:
            old_time = begin_time
        old_progress = 0.0
    if old_progress == new_progress or tot_time == timedelta(0):
        # bad condition
        return None
    速度 = (new_progress - old_progress) / (tot_time / timedelta(hours=1))
    每日用时 = tot_time / workdays(old_time, new_time, worktime)
    return TaskSpeed(速度=速度, 每日用时=每日用时)


def statistic(now_state: State, now_time: datetime) -> StateStats:
    # Goldie点数 = 任务 + 状态 + 其他任务完成
    # 任务
    任务点数 = 0
    任务统计: dict[str, TaskStats] = {}
    总每日用时 = timedelta(0)
    worktime = now_state.工作时段
    for task in now_state.任务.values():
        标记: Literal["*", "-", "=", "!"] = "*"
        if now_time < task.开始:
            标记 = "-"
        elif now_time > task.结束:
            标记 = "!"
        else:
            标记 = "*"
        progress = now_state.进度.get(task.名称, None)
        进度 = 0.0
        进度描述 = None
        用时 = timedelta(0)
        速度 = None
        if progress is not None:
            current = progress[-1]
            进度 = current.进度
            for node in progress:
                if node.描述 is not None or node.进度 != 进度:
                    进度描述 = node.描述
                用时 += node.用时
            速度 = calculate_speed(progress, task.开始, now_time, worktime)
            if task.总数 is not None:
                if 进度 >= task.总数:
                    标记 = "="
        预计 = None
        if 速度 and task.总数 is not None:
            总每日用时 += 速度.每日用时
            预计完成时间 = timedelta(hours=1) * (task.总数 - current.进度) / 速度.速度
            预计可用时间 = workdays(now_time, task.结束, worktime) * 速度.每日用时
            差距 = 预计可用时间 - 预计完成时间
            预计 = TaskEstimate(
                预计完成时间=预计完成时间, 预计可用时间=预计可用时间, 差距=差距
            )
        任务统计[task.名称] = TaskStats(
            进度=进度,
            进度描述=进度描述,
            用时=用时,
            标记=标记,
            速度=速度,
            预计=预计,
        )
    for name, stats in 任务统计.items():
        if stats.标记 == "=":
            end_time = now_state.任务[name].结束
            if end_time > now_time:
                # 提前完成的奖励
                stats.点数 = math.floor(
                    100 * workdays(now_time, end_time, worktime) + 0.5
                )
        else:
            if stats.预计 is not None:
                if stats.预计.差距 > timedelta(0):
                    # 提前的任务，使用推荐用时作为总每日用时来估计“提前天数”，防止“完成后休息”的行为反而提高点数（继续提前完成反而降低点数）
                    stats.点数 = math.floor(stats.预计.差距 / 推荐用时 * 100 + 0.5)
                else:
                    # 落后的任务，使用真正的平均每日用时估计“提前天数”。
                    stats.点数 = math.floor(stats.预计.差距 / 总每日用时 * 100 + 0.5)
            else:
                start_time = now_state.任务[name].开始
                if now_time > start_time:
                    # 延迟开始的惩罚
                    stats.点数 = math.floor(
                        workdays(start_time, now_time, worktime) * -100 + 0.5
                    )
        if stats.点数 is not None:
            任务点数 += stats.点数

    # 状态点数
    状态点数 = 0
    状态生效: list[str] = []
    for status in now_state.状态.values():
        if status.点数 is None:
            continue
        if status.开始 is not None and status.开始 > now_time:
            continue
        if status.结束 is not None and status.结束 < now_time:
            continue
        状态点数 += status.点数
        状态生效.append(status.名称)

    # 其他任务点数
    其他任务点数 = 0
    其他任务生效: list[str] = []
    for todo in now_state.待办事项.values():
        if todo.完成 is not None and todo.完成 > now_time - timedelta(days=3):
            其他任务点数 += todo.点数
            其他任务生效.append(todo.名称)

    Goldie点数 = 任务点数 + 状态点数 + 其他任务点数
    return StateStats(
        Goldie点数=Goldie点数,
        任务点数=任务点数,
        任务统计=任务统计,
        状态点数=状态点数,
        状态生效=状态生效,
        其他任务点数=其他任务点数,
        其他任务生效=其他任务生效,
        总每日用时=总每日用时,
    )


FmtT = TypeVar("FmtT", int, float, datetime, timedelta)


@overload
def fmt(x: FmtT, *, pos: bool = False, timesign: bool = True) -> str: ...


@overload
def fmt(
    x: FmtT,
    y: FmtT,
    *,
    pos: bool = False,
    diff: bool = True,
    timesign: bool = True,
) -> str: ...


def fmt(
    x: FmtT,
    y: FmtT | None = None,
    *,
    pos: bool = False,
    diff: bool = True,
    timesign: bool = True,
) -> str:
    if y is None:
        if isinstance(x, int):
            if pos:
                return f"{x:+d}"
            else:
                return f"{x:d}"
        elif isinstance(x, float):
            if pos:
                return f"{x:+g}"
            else:
                return f"{x:g}"
        elif isinstance(x, datetime):
            if pos:
                raise TypeError("datetime无符号")
            else:
                return x.strftime("%m/%d %H:%M")
        elif isinstance(x, timedelta):
            sign = ""
            if x >= timedelta(0):
                if pos:
                    if timesign:
                        sign = "后"
                    else:
                        sign = "+"
                v = x
            else:
                if pos:
                    if timesign:
                        sign = "前"
                    else:
                        sign = "-"
                v = -x
            if v > timedelta(days=7):
                vstr = f"{v // timedelta(days=1)}天"
            elif v > timedelta(days=1):
                vstr = f"{v // timedelta(days=1)}天{v // timedelta(hours=1) % 24}时"
            elif v > timedelta(hours=1):
                vstr = f"{v // timedelta(hours=1)}时{v // timedelta(minutes=1) % 60}分"
            else:
                vstr = f"{v.seconds // 60}分钟"
            return vstr + sign if timesign else sign + vstr
        else:
            raise TypeError("未知类型")
    else:
        if diff:
            if isinstance(y, timedelta):
                return f"{fmt(y, pos=pos, timesign=timesign)}({fmt(y - x, pos=True, timesign=False)})"
            else:
                return f"{fmt(y, pos=pos, timesign=timesign)}({fmt(y - x, pos=True)})"
        else:
            return f"{fmt(x, pos=pos, timesign=timesign)}→{fmt(y, pos=pos, timesign=timesign)}"


async def index_html(request: aiohttp.web.Request) -> aiohttp.web.FileResponse:
    return aiohttp.web.FileResponse("web/index.html")


def live_server(workbook: str) -> None:
    async def get_points(request: aiohttp.web.Request) -> aiohttp.web.Response:
        data = load_data(workbook)
        now_time = datetime.now()
        now_state = collect_state(data, now_time)
        now_statistic = statistic(now_state, now_time)
        return aiohttp.web.Response(text=fmt(now_statistic.Goldie点数))

    app = aiohttp.web.Application()
    app.add_routes([aiohttp.web.post("/get_points", get_points)])
    app.add_routes([aiohttp.web.get("/", index_html)])
    app.add_routes([aiohttp.web.static("/", "web")])
    aiohttp.web.run_app(app, host="0.0.0.0", port=26019)


@dataclass
class ReportData:
    time: datetime
    state: State
    stats: StateStats


def report_head(
    N: ReportData, P: ReportData | None, *, short: bool = False, diff: bool = True
) -> str:
    report = "blueberry 报告\n"
    if P is not None:
        report += f"时间:{fmt(P.time, N.time, diff=False)}\n"
        if short:
            # 短格式
            report += f"点数:{fmt(P.stats.Goldie点数, N.stats.Goldie点数, diff=diff)} "
            report += f"({fmt(P.stats.任务点数, N.stats.任务点数, diff=diff)},"
            report += f"{fmt(P.stats.状态点数, N.stats.状态点数, diff=diff)},"
            report += f"{fmt(P.stats.其他任务点数, N.stats.其他任务点数, diff=diff)})"
            report += (
                f" 日均:{fmt(P.stats.总每日用时, N.stats.总每日用时, diff=diff)}\n"
            )
        else:
            # 普通格式
            report += (
                f"Goldie点数:{fmt(P.stats.Goldie点数, N.stats.Goldie点数, diff=diff)} "
            )
            report += f"(任务:{fmt(P.stats.任务点数, N.stats.任务点数, diff=diff)}"
            report += f" 状态:{fmt(P.stats.状态点数, N.stats.状态点数, diff=diff)}"
            report += (
                f" 其他:{fmt(P.stats.其他任务点数, N.stats.其他任务点数, diff=diff)})\n"
            )
            report += f"近期平均每日用时:{fmt(P.stats.总每日用时, N.stats.总每日用时, diff=diff)}\n"
    else:
        report += f"时间:{fmt(N.time)}\n"
        if short:
            # 短对比格式
            report += f"点数:{fmt(N.stats.Goldie点数)} "
            report += f"({fmt(N.stats.任务点数)},"
            report += f"{fmt(N.stats.状态点数)},"
            report += f"{fmt(N.stats.其他任务点数)})"
            report += f" 日均:{fmt(N.stats.总每日用时)}\n"
        else:
            # 普通对比格式
            report += f"Goldie点数:{fmt(N.stats.Goldie点数)} "
            report += f"(任务:{fmt(N.stats.任务点数)}"
            report += f" 状态:{fmt(N.stats.状态点数)}"
            report += f" 其他:{fmt(N.stats.其他任务点数)})\n"
            report += f"近期平均每日用时:{fmt(N.stats.总每日用时)}\n"
    return report + "\n"


def report_main_tasks(
    N: ReportData,
    P: ReportData | None,
    *,
    short: bool = False,
    daily: bool = False,
    diff: bool = True,
    change_only: bool = False,
    upcoming: float | None = None,
    verbose: bool = False
) -> str:
    fliter_upcoming: list[str] = []
    fliter_running: list[str] = []  # 包括超时任务
    fliter_running_change: list[str] = []
    fliter_done: list[str] = []
    fliter_done_end_change: list[str] = []
    fliter_done_end: list[str] = []
    for task in N.state.任务.values():
        tstat = N.stats.任务统计[task.名称]
        if task.总数 is not None and tstat.进度 >= task.总数:
            if N.time >= task.结束:
                fliter_done_end.append(task.名称)
                if P is not None and P.time <= task.结束:
                    fliter_done_end_change.append(task.名称)
            else:
                fliter_done.append(task.名称)
        elif tstat.进度 != 0 or N.time >= task.开始:
            fliter_running.append(task.名称)
            if P is not None:
                ptask = P.state.任务.get(task.名称)
                if ptask != task or P.time < task.开始:
                    fliter_running_change.append(task.名称)
                else:
                    pstat = P.stats.任务统计.get(task.名称)
                    if pstat is None or pstat.进度 != tstat.进度:
                        fliter_running_change.append(task.名称)
        else:
            # N.time < task.开始
            if upcoming is None or task.开始 < N.time + timedelta(days=upcoming):
                fliter_upcoming.append(task.名称)

    def report_main_tasks_flitered(title: str, task_names: list[str], *, is_upcoming: bool = False) -> tuple[str, list[list[str]]]:
        if not task_names:
            return "", []
        report = ""
        report += f"{title}:\n"
        table_line:list[list[str]] = []
        for task_name in task_names:
            verbose_str = ""
            task = N.state.任务[task_name]
            tstat = N.stats.任务统计[task_name]
            t点数 = tstat.点数 if tstat.点数 is not None else 0
            statuses = []
            verbose_str += f"  | 开始:{fmt(N.time, task.开始, diff=True)} 结束:{fmt(N.time, task.结束, diff=True)}\n"
            if tstat.速度 is not None:
                verbose_str += f"  | 速度:{fmt(tstat.速度.速度)}/小时 每日平均用时:{fmt(tstat.速度.每日用时)}\n"
            if tstat.预计 is not None:
                verbose_str += f"  | 预计完成时间:{fmt(tstat.预计.预计完成时间)} 预计可用时间:{fmt(tstat.预计.预计可用时间)} 差距:{fmt(tstat.预计.差距)}\n"
            if P is not None:
                ptask = P.state.任务.get(task_name)
                pstat = P.stats.任务统计.get(task_name)
                p进度 = pstat.进度 if pstat is not None else 0
                p用时 = pstat.用时 if pstat is not None else timedelta(0)
                p点数 = pstat.点数 if pstat is not None and pstat.点数 is not None else 0
                point_str = f"{fmt(p点数, t点数, diff=diff)}"
                if ptask != task:
                    statuses.append(f"更新于{fmt(N.time - task.时间)}前")
                if task.总数 is not None:
                    statuses.append(f"{fmt(p进度, tstat.进度, diff=diff)}/{fmt(task.总数)} ({tstat.进度/task.总数:.0%})")
                else:
                    statuses.append(f"{fmt(p进度, tstat.进度, diff=diff)}")
                statuses.append(f"用时{fmt(tstat.用时 - p用时)}")
            else:
                point_str = f"{fmt(t点数)}"
                if task.总数 is not None:
                    statuses.append(f"{fmt(tstat.进度)}/{fmt(task.总数)} ({tstat.进度/task.总数:.0%})")
                else:
                    statuses.append(f"{fmt(tstat.进度)}")
            if is_upcoming:
                statuses.append(f"{fmt(task.开始 - N.time)}后开始")
            elif N.time >= task.结束:
                statuses.append(f"过期{fmt(N.time - task.结束)}")
            else:
                statuses.append(f"剩余{fmt(task.结束 - N.time)}")
            table_line.append([title, point_str, task.标题, *statuses])
            if statuses:
                status_str = "(" +",".join(statuses)+")"
            else:
                status_str = ""
            report += f"- [{point_str}] {task.标题} {status_str}\n"
            if verbose:
                report += verbose_str
            if task.描述 is not None:
                report += indent(task.描述, "  ")
            report += "\n\n"
        return report, table_line


    report_upcoming, table_upcoming = report_main_tasks_flitered("即将开始", fliter_upcoming, is_upcoming=True)
    report_running, table_running = report_main_tasks_flitered("进行中", fliter_running)
    report_running_change, table_running_change = report_main_tasks_flitered("进行中", fliter_running_change)
    report_done_end, table_done_end = report_main_tasks_flitered("已结束", fliter_done_end)
    report_done_end_change, table_done_end_change = report_main_tasks_flitered("已完成", fliter_done_end_change)

    report = ""
    if short:
        table_line:list[list[str]] = []
        if change_only:
            table_line.extend(table_running_change)
            table_line.extend(table_done_end_change)
        else:
            table_line.extend(table_running)
            table_line.extend(table_upcoming)
            table_line.extend(table_done_end)
            table_line.extend(table_done_end_change)
        report += tabulate(table_line, tablefmt="plain") + "\n"
    else:
        if change_only:
            report += report_running_change
            report += report_done_end_change
        else:
            report += report_running
            report += report_upcoming
            report += report_done_end
            report += report_done_end_change
    return report


def main() -> None:

    parser = argparse.ArgumentParser()
    parser.add_argument(
        "-d",
        "--daily",
        action="store_true",
        help="显示当天报告(开始时间将被设为当前时间0点)",
    )
    parser.add_argument(
        "-l", "--live", action="store_true", help="开启实时点数显示HTTP网页"
    )
    parser.add_argument(
        "-f",
        "--from",
        action="store",
        dest="from_",
        help="开始时间(可选, YYYY-MM-DDTHH:MM[:SS])",
    )
    parser.add_argument("-t", "--time", "--to", action="store", help="当前时间")

    parser.add_argument(
        "-w",
        "--workbook",
        action="store",
        default="记录.xlsx",
        help="表格文件路径 (默认: 记录.xlsx)",
    )
    parser.add_argument("-s", "--short", action="store_true", help="简化报告")
    parser.add_argument(
        "-c", "--change-only", action="store_true", help="只显示变化的部分"
    )
    parser.add_argument(
        "-p", "--diff", action="store_true", help="显示 新(±变化) 格式，不用 旧→新 格式"
    )
    parser.add_argument(
        "-n", "--nologging", action="store_true", help="不保存报告为文件"
    )
    parser.add_argument(
        "-v", "--verbose", action="store_true", help="显示详细信息"
    )
    parser.add_argument("-N", "--shownote", action="store_true", help="显示说明")
    parser.add_argument("-D", "--debugspeed", action="store_true", help="显示速度信息")

    args = parser.parse_args()
    print(args)

    if args.live:
        live_server(args.workbook)
        return

    data = load_data(args.workbook)
    tmark = datetime.now().strftime("%Y%m%d-%H%M%S")

    if args.time is not None:
        now_time = datetime.fromisoformat(args.time)
    else:
        now_time = datetime.now()
    yesterday_time = now_time.replace(hour=0, minute=0, second=0)
    if args.from_ is not None:
        prev_time = datetime.fromisoformat(args.from_)
    elif args.daily:
        prev_time = yesterday_time
    else:
        prev_time = None

    now_state = collect_state(data, now_time)
    now_stats = statistic(now_state, now_time)
    now_data = ReportData(now_time, now_state, now_stats)

    yesterday_state = collect_state(data, yesterday_time)
    yesterday_stats = statistic(yesterday_state, yesterday_time)
    yesterday_data = ReportData(yesterday_time, yesterday_state, yesterday_stats)

    prev_data = None
    if prev_time is not None:
        prev_state = collect_state(data, prev_time)
        prev_stats = statistic(prev_state, prev_time)
        prev_data = ReportData(prev_time, prev_state, prev_stats)

    report = ""
    report += report_head(now_data, prev_data, short=args.short, diff=args.diff)

    report += report_main_tasks(
        now_data,
        prev_data,
        short=args.short,
        daily=args.daily,
        diff=args.diff,
        change_only=args.change_only,
        verbose = args.verbose
    )
    # report += report_todo_tasks(
    #     now_data,
    #     prev_data,
    #     change_only=args.change,
    #     upcoming=timedelta,
    #     short=args.short,
    #     diff=args.diff,
    # )
    # report += report_statuses(
    #     now_data, prev_data, change_only=args.change, short=args.short, diff=args.diff
    # )
    # report += report_hints(
    #     now_data, prev_data, change_only=args.change, short=args.short, diff=args.diff
    # )

    if args.shownote:
        report += "-- 说明 --\n"
        with open("blueberry说明.txt", "r", encoding="utf-8") as g:
            report += g.read() + "\n\n"

    if not args.nologging:
        if not os.path.exists("data"):
            os.makedirs("data")
        with open(f"data/{tmark}.json", "w", encoding="utf-8") as f:
            f.write(data.model_dump_json(indent=2))
        with open(f"data/{tmark}.txt", "w", encoding="utf-8") as f:
            f.write(report)

    print(report)


def test_workdays() -> None:
    worktime = [
        WorktimeModel(
            开始=datetime_time(4, 0),
            结束=datetime_time(8, 0),
        ),
        WorktimeModel(
            开始=datetime_time(12, 0),
            结束=datetime_time(16, 0),
        ),
        WorktimeModel(
            开始=datetime_time(22, 0),
            结束=datetime_time(0, 0),
        ),
    ]
    prev_time = datetime.now()
    now_time = prev_time
    for i in range(100):
        now_time += timedelta(hours=0.5)
        days = workdays(prev_time, now_time, worktime)
        print(prev_time, now_time, days)
    for i in range(100):
        prev_time += timedelta(hours=0.5)
        days = workdays(prev_time, now_time, worktime)
        print(prev_time, now_time, days)


if __name__ == "__main__":
    main()
