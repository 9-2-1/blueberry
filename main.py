from typing import Literal, Any, TypeVar, Generic
from dataclasses import dataclass
from datetime import datetime, timedelta, time as datetime_time
from textwrap import indent, wrap
import math
import sys
import os
import warnings
import aiohttp.web

from rich import inspect, traceback as rich_traceback
from rich.pretty import pprint
from pydantic import BaseModel, ConfigDict
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

warnings.simplefilter(action="ignore", category=UserWarning)
rich_traceback.install()


class AppendOnly(BaseModel):
    """
    记录表中共有的两列，一列代表记录的名称，一列代表记录添加、更新或删除的时间。
    """

    名称: str
    时间: datetime


class DeleteModel(AppendOnly):
    """
    代表“删除”的记录行。
    在“时间”删除名为“名称”的行。
    """

    名称: str
    时间: datetime


class TaskModel(AppendOnly):
    model_config = ConfigDict(extra="forbid")
    名称: str
    时间: datetime
    # ---
    标题: str
    描述: str | None = None
    开始: datetime
    结束: datetime
    总数: float | None = None


class ProgressModel(BaseModel):
    model_config = ConfigDict(extra="forbid")
    时间: datetime
    名称: str
    # ---
    进度: float = 0.0
    用时: timedelta = timedelta(0)
    分数: float | None = None
    描述: str | None = None


class StatusModel(AppendOnly):
    model_config = ConfigDict(extra="forbid")
    名称: str
    时间: datetime
    # ---
    标题: str
    描述: str | None = None
    点数: float | None = None  # 点数为None意味着这是一个“取消之前状态”的记录。
    开始: datetime | None = None
    结束: datetime | None = None


class TodoModel(AppendOnly):
    model_config = ConfigDict(extra="forbid")
    名称: str
    时间: datetime
    delete: bool = False
    # ---
    标记: Literal["+", "*", "-"] | None = None  # + 未开始，* 继续，- 等待
    标题: str
    描述: str | None = None
    点数: float = 0.0
    开始: datetime | None = None
    结束: datetime | None = None
    完成: datetime | None = None


class HintModel(BaseModel):
    model_config = ConfigDict(extra="forbid")
    时间: datetime
    # ---
    标题: str
    描述: str | None = None


AppendOnlyModel = TypeVar("AppendOnlyModel", bound=AppendOnly)
T = TypeVar("T")


def parse_table(table: Worksheet) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    keys: list[str] = []
    for row in table.iter_rows(min_row=1):
        for cell in row:
            keys.append(str(cell.value))
    for row in table.iter_rows(min_row=2):
        row_: dict[str, Any] = {}
        for i, cell in enumerate(row):
            if cell.value is None:
                continue
            row_[keys[i]] = cell.value
        if row_:
            rows.append(row_)
    return rows


def parse_model_table(table: Worksheet, datatype: type[T]) -> list[T]:
    parsed_table = parse_table(table)
    ret: list[T] = []
    for x in parsed_table:
        ret.append(datatype(**x))
    return ret


def parse_append_only_table(
    table: Worksheet, datatype: type[AppendOnlyModel]
) -> list[AppendOnlyModel | DeleteModel]:
    parsed_table = parse_table(table)
    ret: list[AppendOnlyModel | DeleteModel] = []
    for x in parsed_table:
        if set(x.keys()) == {"名称", "时间"}:
            ret.append(DeleteModel(**x))
        else:
            ret.append(datatype(**x))
    return ret


class Data(BaseModel):
    任务: list[TaskModel | DeleteModel]
    进度: list[ProgressModel]
    状态: list[StatusModel | DeleteModel]
    待办事项: list[TodoModel | DeleteModel]
    提示: list[HintModel]


class State(BaseModel):
    任务: dict[str, TaskModel]
    进度: dict[str, list[ProgressModel]]
    状态: dict[str, StatusModel]
    待办事项: dict[str, TodoModel]
    提示: list[HintModel]


def mark_delete(x: dict[str, Any]) -> dict[str, Any]:
    if len(set(x.keys()) - {"名称", "时间"}) == 0:
        x["delete"] = True
    return x


def load_data() -> Data:
    wb = load_workbook("记录.xlsx")
    任务 = parse_append_only_table(wb["任务"], TaskModel)
    进度 = parse_model_table(wb["进度"], ProgressModel)
    状态 = parse_append_only_table(wb["状态"], StatusModel)
    待办事项 = parse_append_only_table(wb["待办事项"], TodoModel)
    提示 = parse_model_table(wb["提示"], HintModel)
    return Data(状态=状态, 任务=任务, 进度=进度, 待办事项=待办事项, 提示=提示)


def collect_lines(
    lines: list[AppendOnlyModel | DeleteModel], now_time: datetime
) -> dict[str, AppendOnlyModel]:
    lines = sorted(lines, key=lambda x: x.时间)
    state: dict[str, AppendOnlyModel] = {}
    for line in lines:
        if line.时间 >= now_time:
            continue
        if isinstance(line, DeleteModel):
            del state[line.名称]
        else:
            state[line.名称] = line
    return state


def collect_progress(
    progress: list[ProgressModel], now_time: datetime
) -> dict[str, list[ProgressModel]]:
    progress = sorted(progress, key=lambda x: x.时间)
    state: dict[str, list[ProgressModel]] = {}  # list of progress nodes
    for line in progress:
        if line.时间 >= now_time:
            continue
        if line.名称 in state:
            state[line.名称].append(line)
        else:
            state[line.名称] = [line]
    return state


def collect_hints(lines: list[HintModel], now_time: datetime) -> list[HintModel]:
    lines = sorted(lines, key=lambda x: x.时间)
    return [x for x in lines if x.时间 <= now_time]


def collect_state(data: Data, now_time: datetime) -> State:
    return State(
        任务=collect_lines(data.任务, now_time),
        进度=collect_progress(data.进度, now_time),
        状态=collect_lines(data.状态, now_time),
        待办事项=collect_lines(data.待办事项, now_time),
        提示=collect_hints(data.提示, now_time),
    )


@dataclass
class TaskSpeed:
    速度: float  # per hour
    每日用时: timedelta


@dataclass
class TaskEstimate:
    预计完成时间: timedelta
    预计可用时间: timedelta
    差距: timedelta


@dataclass
class TaskStats:
    进度: float
    标记: Literal["*", "-", "=", "!"]
    # * 进行中 - 等待开始 = 已完成 ! 已超时
    速度: TaskSpeed | None = None  # 开始的任务才能计算速度
    预计: TaskEstimate | None = None  # 开始但未完成的任务才能估计完成时间
    点数: float | None = (
        None  # 点数。到时未开始: -100 * 延后天数。已开始：100 * (时间差距 / 每日用时)。已结束：100 * 剩余天数
    )


@dataclass
class StateStats:
    Goldie点数: float
    任务点数: float
    任务统计: dict[str, TaskStats]
    状态点数: float
    状态生效: list[str]
    其他任务点数: float
    其他任务生效: list[str]
    总每日用时: timedelta


def calculate_speed(
    progress: list[ProgressModel], begin_time: datetime, now_time: datetime
) -> TaskSpeed | None:
    # 近期记录
    MIN_TIME = timedelta(hours=2)
    MIN_TIMESPAN = timedelta(days=3)
    new_time = now_time
    tot_time = timedelta(0)
    new_progress = progress[-1].进度
    found = False
    count = 0
    # 这里需要注意时间的计算方式：
    # 进度记录描述的是 “花费‘用时’时间后，在‘时间’让进度达到了‘进度’”。
    # 选择“开始的记录”后“开始的记录”本身的时间不包含在总时间内（那是起点）
    for p in reversed(progress):
        old_time = p.时间
        if tot_time > MIN_TIME and (new_time - old_time) > MIN_TIMESPAN:
            found = True
            old_progress = p.进度
            break
        tot_time += p.用时
        if p.用时 != timedelta(0):
            count += 1
    if not found:
        old_time = begin_time
        old_progress = 0.0
    if old_progress == new_progress or tot_time == timedelta(0):
        # bad condition
        return None
    速度 = (new_progress - old_progress) / (tot_time / timedelta(hours=1))
    每日用时 = tot_time / ((new_time - old_time) / timedelta(days=1))
    return TaskSpeed(速度=速度, 每日用时=每日用时)


def statistic(now_state: State, now_time: datetime) -> StateStats:
    # Goldie点数 = 任务 + 状态 + 其他任务完成
    # 任务

    任务点数 = 0.0
    任务统计: dict[str, TaskStats] = {}
    总每日用时 = timedelta(0)
    for task in now_state.任务.values():
        标记: Literal["*", "-", "=", "!"] = "*"
        if now_time < task.开始:
            标记 = "-"
        elif now_time > task.结束:
            标记 = "!"
        else:
            标记 = "*"
        progress = now_state.进度.get(task.名称, None)
        进度 = 0.0
        速度 = None
        if progress is not None:
            current = progress[-1]
            进度 = current.进度
            速度 = calculate_speed(progress, task.开始, now_time)
            if task.总数 is not None:
                if 进度 >= task.总数:
                    标记 = "="
        预计 = None
        if 速度 and task.总数 is not None:
            总每日用时 += 速度.每日用时
            预计完成时间 = timedelta(hours=1) * (task.总数 - current.进度) / 速度.速度
            预计可用时间 = (task.结束 - now_time) / timedelta(days=1) * 速度.每日用时
            差距 = 预计可用时间 - 预计完成时间
            预计 = TaskEstimate(
                预计完成时间=预计完成时间, 预计可用时间=预计可用时间, 差距=差距
            )
        任务统计[task.名称] = TaskStats(
            进度=进度,
            标记=标记,
            速度=速度,
            预计=预计,
        )
    for name, stats in 任务统计.items():
        if stats.标记 == "=":
            end_time = now_state.任务[name].结束
            if end_time > now_time:
                # 提前完成的奖励
                stats.点数 = 100 * (end_time - now_time) / timedelta(days=1)
        else:
            if stats.预计 is not None:
                推荐用时 = timedelta(hours=6.5)
                if stats.预计.差距 > timedelta(0):
                    # 提前的任务，使用推荐用时作为总每日用时来估计“提前天数”，防止“完成后休息”的行为反而提高点数（继续提前完成反而降低点数）
                    stats.点数 = stats.预计.差距 / 总每日用时 * 100
                else:
                    # 落后的任务，使用真正的平均每日用时估计“提前天数”。
                    stats.点数 = stats.预计.差距 / 总每日用时 * 100
            else:
                start_time = now_state.任务[name].开始
                if now_time > start_time:
                    # 延迟开始的惩罚
                    stats.点数 = -100 * (now_time - start_time) / timedelta(days=1)
        if stats.点数 is not None:
            任务点数 += stats.点数

    # 状态点数
    状态点数 = 0.0
    状态生效: list[str] = []
    for status in now_state.状态.values():
        if status.点数 is None:
            continue
        if status.开始 is not None and status.开始 > now_time:
            continue
        if status.结束 is not None and status.结束 < now_time:
            continue
        状态点数 += status.点数
        状态生效.append(status.名称)

    # 其他任务点数
    其他任务点数 = 0.0
    其他任务生效: list[str] = []
    for todo in now_state.待办事项.values():
        if todo.完成 is not None and todo.完成 > now_time - timedelta(days=3):
            其他任务点数 += todo.点数
            其他任务生效.append(todo.名称)

    Goldie点数 = 任务点数 + 状态点数 + 其他任务点数
    return StateStats(
        Goldie点数=Goldie点数,
        任务点数=任务点数,
        任务统计=任务统计,
        状态点数=状态点数,
        状态生效=状态生效,
        其他任务点数=其他任务点数,
        其他任务生效=其他任务生效,
        总每日用时=总每日用时,
    )


def deltafmt_unsigned(delta: timedelta) -> str:
    if delta > timedelta(days=7):
        return f"{delta // timedelta(days=1)}d"
    elif delta > timedelta(days=1):
        return f"{delta // timedelta(days=1)}d{delta // timedelta(hours=1) % 24}h"
    elif delta > timedelta(hours=1):
        return f"{delta // timedelta(hours=1)}h{delta // timedelta(minutes=1) % 60}m"
    else:
        return f"{delta.seconds // 60}min"


def deltafmt_signed(
    delta: timedelta, positive: bool = False, now_str: str = "0"
) -> str:
    if delta > timedelta(minutes=1):
        if positive:
            return f"+{deltafmt_unsigned(delta)}"
        else:
            return deltafmt_unsigned(delta)
    elif delta < timedelta(minutes=-1):
        return f"-{deltafmt_unsigned(-delta)}"
    else:
        return now_str


def timefmt(t: datetime, now: datetime | None = None) -> str:
    timestr = t.strftime("%m/%d %H:%M")
    if now is not None:
        delta = t - now
        timestr += f" ({deltafmt_signed(delta, True, "现在")})"
    return timestr


def pointfmt(p: float) -> str:
    p_i = math.floor(p + 0.5)
    if p_i == 0:
        return "0"
    else:
        return f"{p_i:+g}"


def change(x: str, y: str) -> str:
    if x == y:
        return x
    else:
        return f"{x} -> {y}"


def boardprint(number: float) -> None:
    print(f"\r{pointfmt(number)}", end="")


async def index_html(request: aiohttp.web.Request) -> aiohttp.web.FileResponse:
    return aiohttp.web.FileResponse("web/index.html")


async def get_points(request: aiohttp.web.Request) -> aiohttp.web.Response:
    data = load_data()
    now_time = datetime.now()
    now_state = collect_state(data, now_time)
    now_statistic = statistic(now_state, now_time)
    return aiohttp.web.Response(text=pointfmt(now_statistic.Goldie点数))


def live_server() -> None:
    app = aiohttp.web.Application()
    app.add_routes([aiohttp.web.post("/get_points", get_points)])
    app.add_routes([aiohttp.web.get("/", index_html)])
    app.add_routes([aiohttp.web.static("/", "web")])
    aiohttp.web.run_app(app, host="0.0.0.0", port=26019)


def main() -> None:

    if len(sys.argv) > 1:
        if sys.argv[1] == "live":
            live_server()
            return

    data = load_data()
    tmark = datetime.now().strftime("%Y%m%d-%H%M%S")

    if len(sys.argv) > 2:
        prev_time = datetime.fromisoformat(sys.argv[1])
        now_time = datetime.fromisoformat(sys.argv[2])
    elif len(sys.argv) > 1:
        prev_time = datetime.fromisoformat(sys.argv[1])
        now_time = datetime.fromisoformat(sys.argv[1])
    else:
        now_time = datetime.now()
        prev_time = now_time.replace(hour=0, minute=0, second=0)

    now_state = collect_state(data, now_time)
    prev_state = collect_state(data, prev_time)

    now_statistic = statistic(now_state, now_time)
    prev_statistic = statistic(prev_state, prev_time)

    with open(f"data/{tmark}.json", "w", encoding="utf-8") as f:
        f.write(data.model_dump_json(indent=2))

    mark: str
    with open(f"data/{tmark}.txt", "w", encoding="utf-8") as f:

        def write(s: str) -> None:
            print(s)
            print(s, file=f)

        write("-- Blueberry --")
        write(
            f"时间: {change(timefmt(prev_time), timefmt(now_time))} ({deltafmt_signed(now_time-prev_time, True, '0')})"
        )
        write(
            f"Goldie点数: {change(pointfmt(prev_statistic.Goldie点数), pointfmt(now_statistic.Goldie点数))}"
        )
        write(
            f"- 主要: {change(pointfmt(prev_statistic.任务点数), pointfmt(now_statistic.任务点数))}"
        )
        write(
            f"- 状态: {change(pointfmt(prev_statistic.状态点数), pointfmt(now_statistic.状态点数))}"
        )
        write(
            f"- 其他任务: {change(pointfmt(prev_statistic.其他任务点数), pointfmt(now_statistic.其他任务点数))}"
        )
        write("")

        write("-- 主要任务 --")
        # 8小时(时长) * 80%(工作-休息比) ≈ 6.5小时
        推荐用时 = timedelta(hours=6.5)
        write(
            f"平均每日: {change(deltafmt_signed(prev_statistic.总每日用时), deltafmt_signed(now_statistic.总每日用时))} (推荐用时的 {now_statistic.总每日用时 / 推荐用时 :.0%})"
        )
        for task in now_state.任务.values():
            tstats = now_statistic.任务统计.get(task.名称, None)
            assert tstats is not None
            mark = tstats.标记
            点数str = ""
            if tstats.点数 is not None:
                prev_stat = prev_statistic.任务统计.get(task.名称)
                prev_point = prev_stat.点数 if prev_stat is not None else None
                if prev_point is not None:
                    点数str = f" [{change(pointfmt(prev_point),pointfmt(tstats.点数))}]"
                else:
                    点数str = f" [{pointfmt(tstats.点数)}]"
            write(
                f"{mark}{点数str} {task.标题} 开始: {timefmt(task.开始, now_time)} 结束: {timefmt(task.结束, now_time)}"
            )
            if task.总数 is not None:
                write(
                    f"  > 已完成: {tstats.进度:g}/{task.总数:g} ({tstats.进度/task.总数:.0%})"
                )
            else:
                write(f"  > 已完成: {tstats.进度:g}")
            if tstats.速度 is not None:
                write(
                    f"  > 速度: {tstats.速度.速度:.3g}/h 平均每日用时: {deltafmt_signed(tstats.速度.每日用时)}"
                )
            if tstats.预计 is not None:
                write(
                    f"  > 预计完成时间: {deltafmt_signed(tstats.预计.预计完成时间)} 预计可用时间: {deltafmt_signed(tstats.预计.预计可用时间)} 差距: {deltafmt_signed(tstats.预计.差距, True)}"
                )
            if task.描述 is not None:
                write(indent(task.描述, "    "))
            write("")

        write("-- 其他任务 --")
        for mark, title in [
            ("*", "继续"),
            ("+", "可以进行"),
            ("-", "等待"),
        ]:
            todos = [
                x
                for x in now_state.待办事项.values()
                if x.标记 == mark and x.完成 is None
            ]
            if todos:
                write(f"{title}:")
                todos = sorted(
                    todos,
                    key=lambda x: (
                        x.开始
                        if x.开始 is not None
                        else x.结束 if x.结束 is not None else datetime.max
                    ),
                )
                for todo in todos:
                    write(
                        f"{mark} (预计{pointfmt(todo.点数)}) {todo.标题}"
                        + (
                            " 开始: " + timefmt(todo.开始, now_time)
                            if todo.开始 is not None
                            else ""
                        )
                        + (
                            " 结束: " + timefmt(todo.结束, now_time)
                            if todo.结束 is not None
                            else ""
                        )
                    )
                    if todo.描述 is not None:
                        write(indent(todo.描述, "    "))
                    write("")
        if now_statistic.其他任务生效:
            write("最近完成:")
            for todo_name in now_statistic.其他任务生效:
                todo = now_state.待办事项[todo_name]
                write(
                    f"= [{pointfmt(todo.点数)}] {todo.标题}"
                    + (
                        " 开始: " + timefmt(todo.开始, now_time)
                        if todo.开始 is not None
                        else ""
                    )
                    + (
                        " 结束: " + timefmt(todo.结束, now_time)
                        if todo.结束 is not None
                        else ""
                    )
                    + (
                        " 完成: " + timefmt(todo.完成, now_time)
                        if todo.完成 is not None
                        else ""
                    )
                )
                if todo.描述 is not None:
                    write(indent(todo.描述, "    "))
                write("")

        write("-- 提示 --")
        hints = reversed(now_state.提示)
        # 只显示最近一段时间或者最近几条提示
        MAX_HINTS = 5
        MAX_HINTS_TIME = timedelta(days=1)
        hints_n = 0
        for hint in hints:
            if hints_n >= MAX_HINTS and hint.时间 < now_time - MAX_HINTS_TIME:
                break
            hints_n += 1
            write(f"o {hint.标题} {timefmt(hint.时间, now_time)}")
            if hint.描述 is not None:
                write(indent(hint.描述, "    "))
            write("")
        write("")

        write("-- 状态 --")
        for status in sorted(
            now_state.状态.values(), key=lambda x: x.时间, reverse=True
        ):
            active = True
            if status.点数 is None:
                active = False
            if status.开始 is not None and status.开始 > now_time:
                active = False
            if status.结束 is not None and status.结束 < now_time:
                active = False
            if prev_state.状态.get(status.名称) == status:
                # 状态发生变化的时候不隐藏
                if not active:
                    continue
            if status.点数 is None:
                mark = "="
            elif status.点数 > 0:
                mark = "+"
            elif status.点数 < 0:
                mark = "-"
            else:
                mark = "o"
            pointstr = ""
            if status.点数 is not None:
                if active:
                    pointstr = f" [{pointfmt(status.点数)}]"
            write(
                f"{mark}{pointstr} {status.标题}"
                + (
                    " 开始: " + timefmt(status.开始, now_time)
                    if status.开始 is not None
                    else ""
                )
                + (
                    " 结束: " + timefmt(status.结束, now_time)
                    if status.结束 is not None
                    else ""
                )
            )
            if status.描述 is not None:
                write(indent(status.描述, "    "))
        write("")

        write("-- 说明 --")
        with open("blueberry说明.txt", "r", encoding="utf-8") as g:
            write(g.read())

    os.startfile(f"data\\{tmark}.txt")


if __name__ == "__main__":
    main()
